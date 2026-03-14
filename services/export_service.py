"""Excel ve CSV export servisi."""
import csv, logging
from pathlib import Path
from typing import List
from constants import SYM_MAP

logger = logging.getLogger("export_service")

HEADERS = ["Teklif No", "Firma", "Tarih", "Para Birimi", "Toplam Tutar",
           "Durum", "Vade", "Ödeme", "İlgili Kişi"]

def _row(o: dict) -> list:
    sym = SYM_MAP.get(o.get("currency", ""), "")
    return [
        o.get("offer_no", ""),
        o.get("company_name", ""),
        o.get("date", ""),
        o.get("currency", ""),
        f"{o.get('total_amount', 0):,.2f} {sym}".strip(),
        o.get("status", ""),
        o.get("validity", ""),
        o.get("payment_term", ""),
        o.get("contact_person", ""),
    ]


def export_excel(offers: List[dict], path: str) -> str:
    """Verilen teklif listesini Excel'e yazar. path döner."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl kurulu değil: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Teklifler"

    # Başlık satırı
    header_fill = PatternFill("solid", fgColor="1E4D8C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 20

    # Veri satırları
    alt_fill = PatternFill("solid", fgColor="F0F4FA")
    for r_idx, o in enumerate(offers, 2):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c_idx, val in enumerate(_row(o), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r_idx].height = 16

    # Sütun genişlikleri
    widths = [18, 28, 12, 10, 16, 12, 10, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Özet satırı
    sum_row = len(offers) + 2
    ws.cell(row=sum_row, column=1, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=sum_row, column=5,
            value=sum(o.get("total_amount", 0) for o in offers)).font = Font(bold=True)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Excel export: %s (%d teklif)", path, len(offers))
    return path


def export_csv(offers: List[dict], path: str) -> str:
    """Verilen teklif listesini CSV'ye yazar. path döner."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(HEADERS)
        for o in offers:
            w.writerow(_row(o))
    logger.info("CSV export: %s (%d teklif)", path, len(offers))
    return path
